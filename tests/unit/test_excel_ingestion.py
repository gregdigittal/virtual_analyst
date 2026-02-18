"""Unit tests for Excel ingestion orchestrator: start_ingestion, create_draft_from_mapping, mapping conversion."""

from __future__ import annotations

from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from apps.api.app.services.excel_ingestion import (
    create_draft_from_mapping,
    start_ingestion,
)
from apps.api.app.services.excel_parser import parse_workbook


def _minimal_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "X"
    ws["B1"] = "Y"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_start_ingestion_rejects_non_xlsx() -> None:
    store = MagicMock()
    conn = MagicMock()
    conn.execute = AsyncMock()
    with pytest.raises(ValueError, match="Only .xlsx"):
        await start_ingestion("t1", "u1", "file.csv", b"data", store, conn)


@pytest.mark.asyncio
async def test_start_ingestion_rejects_over_10mb() -> None:
    store = MagicMock()
    conn = MagicMock()
    conn.execute = AsyncMock()
    data = _minimal_xlsx_bytes()
    huge = data + b"x" * (11 * 1024 * 1024 - len(data))
    with pytest.raises(ValueError, match="exceeds"):
        await start_ingestion("t1", "u1", "file.xlsx", huge, store, conn)


@pytest.mark.asyncio
async def test_start_ingestion_returns_ingestion_id() -> None:
    store = MagicMock()
    conn = MagicMock()
    conn.execute = AsyncMock()
    data = _minimal_xlsx_bytes()
    ingestion_id = await start_ingestion("t1", "u1", "test.xlsx", data, store, conn)
    assert ingestion_id.startswith("xi_")
    assert len(ingestion_id) == 15  # xi_ + 12 hex
    store.save.assert_called_once()
    conn.execute.assert_called_once()
    call_args = conn.execute.call_args[0]
    assert call_args[1] == "t1"  # tenant_id
    assert call_args[2] == ingestion_id
    assert call_args[3] == "test.xlsx"
    assert call_args[4] == len(data)
    assert "uploaded" in str(conn.execute.call_args)


@pytest.mark.asyncio
async def test_create_draft_from_mapping_populates_workspace() -> None:
    mapping = {
        "metadata": {"entity_name": "Test Co", "currency": "USD", "horizon_months": 24},
        "revenue_streams": [
            {"label": "Product", "stream_type": "unit_sale", "source_sheet": "Rev", "source_row_label": "B5"},
        ],
        "cost_items": [
            {"label": "COGS", "category": "cogs", "source_sheet": "P&L"},
        ],
        "capex_items": [
            {"label": "Equipment", "amount": 10000, "month": 0, "useful_life_months": 60},
        ],
        "working_capital": {},
        "funding": {},
        "unmapped_items": [{"label": "Other", "source_sheet": "Misc", "reason": "No mapping"}],
    }
    store = MagicMock()
    conn = MagicMock()
    conn.execute = AsyncMock()
    conn.transaction = MagicMock()
    conn.transaction.return_value.__aenter__ = AsyncMock(return_value=None)
    conn.transaction.return_value.__aexit__ = AsyncMock(return_value=None)

    with patch("apps.api.app.services.excel_ingestion.ensure_tenant", new_callable=AsyncMock), \
         patch("apps.api.app.services.excel_ingestion.create_audit_event", new_callable=AsyncMock):
        draft_id = await create_draft_from_mapping(
            "t1", "u1", "xi_abc123", mapping, "model.xlsx", store, conn
        )
    assert draft_id.startswith("ds_")
    store.save.assert_called_once()
    call_args = store.save.call_args[0]
    assert call_args[0] == "t1"
    assert call_args[1] == "draft_workspace"
    assert call_args[2] == draft_id
    workspace = call_args[3]
    assert workspace["assumptions"]["revenue_streams"]
    assert len(workspace["assumptions"]["revenue_streams"]) == 1
    assert workspace["assumptions"]["revenue_streams"][0]["label"] == "Product"
    assert "Excel: model.xlsx" in workspace["assumptions"]["revenue_streams"][0]["source"]
    assert len(workspace["assumptions"]["cost_structure"]["variable_costs"]) == 1
    assert workspace["assumptions"]["cost_structure"]["variable_costs"][0]["label"] == "COGS"
    assert workspace["assumptions"]["capex"]["items"]
    assert workspace["assumptions"]["capex"]["items"][0]["label"] == "Equipment"
    assert workspace["assumptions"]["capex"]["items"][0]["amount"] == 10000
    assert workspace["custom"] == mapping["unmapped_items"]
    assert any(e.get("proposed_by") == "llm" for e in workspace["evidence"])


@pytest.mark.asyncio
async def test_create_draft_from_mapping_empty_mapping() -> None:
    mapping = {
        "metadata": {},
        "revenue_streams": [],
        "cost_items": [],
        "capex_items": [],
        "working_capital": {},
        "funding": {},
        "unmapped_items": [],
    }
    store = MagicMock()
    conn = MagicMock()
    conn.execute = AsyncMock()
    conn.transaction = MagicMock()
    conn.transaction.return_value.__aenter__ = AsyncMock(return_value=None)
    conn.transaction.return_value.__aexit__ = AsyncMock(return_value=None)
    with patch("apps.api.app.services.excel_ingestion.ensure_tenant", new_callable=AsyncMock), \
         patch("apps.api.app.services.excel_ingestion.create_audit_event", new_callable=AsyncMock):
        draft_id = await create_draft_from_mapping(
            "t1", None, "xi_xyz", mapping, "empty.xlsx", store, conn
        )
    assert draft_id.startswith("ds_")
    workspace = store.save.call_args[0][3]
    assert workspace["assumptions"]["revenue_streams"] == []
    assert workspace["assumptions"]["cost_structure"]["variable_costs"] == []
    assert workspace["assumptions"]["cost_structure"]["fixed_costs"] == []
    assert workspace["assumptions"]["capex"]["items"] == []
    assert workspace["custom"] == []
