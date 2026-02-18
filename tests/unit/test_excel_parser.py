"""Unit tests for Excel parser: parse_workbook, classify_sheets, extract_assumption_candidates."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.api.app.services.excel_parser import (
    ExcelParseResult,
    SheetInfo,
    classify_sheets,
    extract_assumption_candidates,
    parse_workbook,
)


def _make_small_xlsx() -> bytes:
    """Minimal .xlsx: 2 sheets, headers, values, one formula."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Assumptions"
    ws["A1"] = "Item"
    ws["B1"] = "Value"
    ws["A2"] = "Revenue growth"
    ws["B2"] = 0.05
    ws["A3"] = "Horizon"
    ws["B3"] = 12
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Jan"
    ws2["B1"] = "Feb"
    ws2["A2"] = 100
    ws2["B2"] = "=A2*1.1"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_workbook_returns_structure() -> None:
    data = _make_small_xlsx()
    result = parse_workbook(data, filename="test.xlsx")
    assert isinstance(result, ExcelParseResult)
    assert result.filename == "test.xlsx"
    assert result.file_size_bytes == len(data)
    assert result.sheet_count == 2
    assert len(result.sheets) == 2
    names = {s.name for s in result.sheets}
    assert "Assumptions" in names
    assert "Summary" in names
    assert result.total_formulas >= 1
    assert result.total_cross_refs >= 0
    assert result.dependency_graph is not None


def test_parse_workbook_sheet_info() -> None:
    data = _make_small_xlsx()
    result = parse_workbook(data)
    for sheet in result.sheets:
        assert isinstance(sheet, SheetInfo)
        assert sheet.name
        assert sheet.dimensions
        assert sheet.row_count >= 0
        assert sheet.col_count >= 0
        assert isinstance(sheet.headers, list)
        assert isinstance(sheet.sample_rows, list)
        assert isinstance(sheet.formula_patterns, list)
        assert isinstance(sheet.referenced_sheets, list)


def test_classify_sheets_heuristic() -> None:
    data = _make_small_xlsx()
    result = parse_workbook(data)
    classification = classify_sheets(result)
    assert isinstance(classification, dict)
    for name in [s.name for s in result.sheets]:
        assert name in classification
        assert classification[name] in (
            "financial",
            "documentation",
            "project_management",
            "data_reference",
            "empty",
            "unknown",
        )


def test_classify_sheets_financial_like() -> None:
    """Sheet with month-like headers and formulas is classified financial or unknown."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "FM"
    for c, h in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun"], start=1):
        ws.cell(row=1, column=c, value=h)
    for r in range(2, 8):
        for c in range(1, 7):
            ws.cell(row=r, column=c, value=f"=B1*{r}")
    buf = BytesIO()
    wb.save(buf)
    result = parse_workbook(buf.getvalue())
    classification = classify_sheets(result)
    assert classification.get("FM") in ("financial", "unknown")


def test_extract_assumption_candidates() -> None:
    data = _make_small_xlsx()
    result = parse_workbook(data)
    assumptions_sheet = next(s for s in result.sheets if s.name == "Assumptions")
    candidates = extract_assumption_candidates(assumptions_sheet)
    assert isinstance(candidates, list)
    assert len(candidates) >= 2
    for c in candidates:
        assert "row_index" in c
        assert "label" in c
        assert "value" in c
        assert "column" in c
        assert "data_type_guess" in c


def test_parse_workbook_rejects_corrupt_bytes() -> None:
    with pytest.raises(ValueError, match="Invalid|corrupt"):
        parse_workbook(b"not a zip file", filename="bad.xlsx")


def test_parse_workbook_rejects_empty() -> None:
    # Empty bytes are not valid xlsx (zip)
    with pytest.raises(ValueError):
        parse_workbook(b"", filename="empty.xlsx")


def test_file_size_limit_not_in_parser() -> None:
    """Parser does not enforce 10MB; that is in the API. Parser only enforces timeout/size caps per sheet."""
    data = _make_small_xlsx()
    result = parse_workbook(data)
    assert result.file_size_bytes == len(data)
