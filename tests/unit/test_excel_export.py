"""Unit tests for excel_export.py — XLSX workbook generation."""

from __future__ import annotations

import pytest
from io import BytesIO

from openpyxl import load_workbook

from apps.api.app.services.excel_export import build_run_excel


def _sample_statements() -> dict:
    return {
        "income_statement": [
            {"revenue": 100_000, "cogs": 40_000, "ebitda": 60_000},
            {"revenue": 110_000, "cogs": 44_000, "ebitda": 66_000},
        ],
        "balance_sheet": [
            {"total_assets": 500_000, "total_equity": 300_000},
        ],
        "cash_flow": [
            {"operating_cf": 50_000},
        ],
    }


def test_build_run_excel_creates_three_sheets() -> None:
    """Output should contain Income Statement, Balance Sheet, Cash Flow sheets."""
    xlsx_bytes = build_run_excel(_sample_statements())
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "Income Statement" in wb.sheetnames
    assert "Balance Sheet" in wb.sheetnames
    assert "Cash Flow" in wb.sheetnames
    wb.close()


def test_build_run_excel_income_statement_values() -> None:
    """IS sheet should have line items as rows and period values in columns."""
    xlsx_bytes = build_run_excel(_sample_statements())
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb["Income Statement"]
    # Row 1 = header, Row 2 = first line item ("Revenue"), Col 2 = P0 value
    assert ws.cell(row=2, column=2).value == 100_000
    assert ws.cell(row=2, column=3).value == 110_000
    wb.close()


def test_build_run_excel_with_kpis_sheet() -> None:
    """When KPIs are passed, a KPIs sheet should be added."""
    kpis = [
        {"period_index": 0, "gross_margin": 0.60, "net_margin": 0.45},
        {"period_index": 1, "gross_margin": 0.62, "net_margin": 0.47},
    ]
    xlsx_bytes = build_run_excel(_sample_statements(), kpis=kpis)
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "KPIs" in wb.sheetnames
    wb.close()


def test_build_run_excel_empty_statements_raises() -> None:
    """build_run_excel with all-empty lists raises (workbook has no visible sheets)."""
    with pytest.raises(Exception):
        build_run_excel({"income_statement": [], "balance_sheet": [], "cash_flow": []})
