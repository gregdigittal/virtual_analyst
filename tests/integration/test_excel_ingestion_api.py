"""Integration tests for Excel ingestion API: validation, RLS, RBAC."""

from __future__ import annotations

from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook

from apps.api.app.deps import get_artifact_store, get_llm_router
from apps.api.app.main import app
from shared.fm_shared.storage import ArtifactStore

# Patch target: tenant_conn is used inside the router when handling requests
TENANT_CONN_PATCH = "apps.api.app.routers.excel_ingestion.tenant_conn"


def _mock_tenant_conn_context(fetchrow_result=None, fetch_result=None, execute_result="DELETE 1"):
    """Build a mock async context manager for tenant_conn that yields a mock connection."""
    conn = MagicMock()
    conn.fetchrow = AsyncMock(return_value=fetchrow_result)
    conn.fetch = AsyncMock(return_value=fetch_result if fetch_result is not None else [])
    conn.execute = AsyncMock(return_value=execute_result)
    cm = MagicMock()
    cm.__aenter__ = AsyncMock(return_value=conn)
    cm.__aexit__ = AsyncMock(return_value=None)
    return cm


def _minimal_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Item"
    ws["B1"] = "Value"
    ws["A2"] = 1
    ws["B2"] = 2
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def in_memory_store() -> ArtifactStore:
    return ArtifactStore(supabase_client=None)


@pytest.fixture
def mock_llm() -> MagicMock:
    m = MagicMock()
    m.complete_with_routing = AsyncMock(return_value=MagicMock(
        content={
            "sheets": [
                {"sheet_name": "Sheet1", "classification": "assumptions", "role": "Inputs", "confidence": "high", "is_financial_core": True},
            ],
            "model_summary": {
                "entity_name": "Test", "industry": "Tech", "model_type": "startup",
                "currency_guess": "USD", "horizon_months_guess": 12,
            },
        },
        raw_text="{}",
    ))
    return m


@pytest.fixture
def client(in_memory_store: ArtifactStore, mock_llm: MagicMock) -> AsyncClient:
    app.dependency_overrides[get_artifact_store] = lambda: in_memory_store
    app.dependency_overrides[get_llm_router] = lambda: mock_llm
    try:
        transport = ASGITransport(app=app)
        yield AsyncClient(transport=transport, base_url="http://test")
    finally:
        app.dependency_overrides.pop(get_artifact_store, None)
        app.dependency_overrides.pop(get_llm_router, None)


@pytest.mark.asyncio
async def test_upload_requires_x_tenant_id(client: AsyncClient) -> None:
    data = _minimal_xlsx_bytes()
    r = await client.post(
        "/api/v1/excel-ingestion/upload",
        files={"file": ("test.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 422 or r.status_code == 400  # FastAPI may return 422 for missing header


@pytest.mark.asyncio
async def test_get_ingestion_requires_x_tenant_id(client: AsyncClient) -> None:
    r = await client.get("/api/v1/excel-ingestion/xi_abc123")
    assert r.status_code in (400, 422)


@pytest.mark.asyncio
async def test_list_ingestions_requires_x_tenant_id(client: AsyncClient) -> None:
    r = await client.get("/api/v1/excel-ingestion")
    assert r.status_code in (400, 422)


@patch(TENANT_CONN_PATCH)
@pytest.mark.asyncio
async def test_get_ingestion_404_unknown(mock_tenant_conn: MagicMock, client: AsyncClient) -> None:
    mock_tenant_conn.return_value = _mock_tenant_conn_context(fetchrow_result=None)
    r = await client.get(
        "/api/v1/excel-ingestion/xi_nonexistent123",
        headers={"X-Tenant-ID": "t1"},
    )
    assert r.status_code == 404


@patch(TENANT_CONN_PATCH)
@pytest.mark.asyncio
async def test_analyze_404_unknown(mock_tenant_conn: MagicMock, client: AsyncClient) -> None:
    mock_tenant_conn.return_value = _mock_tenant_conn_context(fetchrow_result=None)
    r = await client.post(
        "/api/v1/excel-ingestion/xi_nonexistent123/analyze",
        headers={"X-Tenant-ID": "t1"},
    )
    assert r.status_code == 404


@patch(TENANT_CONN_PATCH)
@pytest.mark.asyncio
async def test_create_draft_404_unknown(mock_tenant_conn: MagicMock, client: AsyncClient) -> None:
    mock_tenant_conn.return_value = _mock_tenant_conn_context(fetchrow_result=None)
    r = await client.post(
        "/api/v1/excel-ingestion/xi_nonexistent123/create-draft",
        headers={"X-Tenant-ID": "t1", "X-User-ID": "u1"},
    )
    assert r.status_code == 404


@patch(TENANT_CONN_PATCH)
@pytest.mark.asyncio
async def test_delete_ingestion_404_unknown(mock_tenant_conn: MagicMock, client: AsyncClient) -> None:
    mock_tenant_conn.return_value = _mock_tenant_conn_context(execute_result="DELETE 0")
    r = await client.delete(
        "/api/v1/excel-ingestion/xi_nonexistent123",
        headers={"X-Tenant-ID": "t1"},
    )
    assert r.status_code == 404
