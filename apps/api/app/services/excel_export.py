"""Excel export: build .xlsx with IS/BS/CF sheets from run statements (VA-P5-01)."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook


def _sheet_from_period_dicts(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    skip_keys: frozenset[str] = frozenset({"period_index"}),
) -> None:
    """Write one sheet: row 1 = headers (Period, then period indices), rows 2+ = line item, values per period."""
    if not rows:
        return
    ws = wb.create_sheet(title=sheet_name[:31])
    keys = [k for k in rows[0].keys() if k not in skip_keys]
    n_periods = len(rows)
    ws.cell(row=1, column=1, value="Line Item")
    for c in range(n_periods):
        ws.cell(row=1, column=c + 2, value=f"P{c}" if c < n_periods else "")
    for r, key in enumerate(keys, start=2):
        ws.cell(row=r, column=1, value=key.replace("_", " ").title())
        for c, period_row in enumerate(rows):
            val = period_row.get(key)
            if isinstance(val, (int, float)):
                ws.cell(row=r, column=c + 2, value=val)
            else:
                ws.cell(row=r, column=c + 2, value=val)


def build_run_excel(
    statements: dict[str, Any],
    kpis: list[dict[str, Any]] | None = None,
) -> bytes:
    """
    Build an Excel workbook with Income Statement, Balance Sheet, Cash Flow sheets.
    Optionally add a KPIs sheet (one row per period, columns = KPI keys).
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    is_list = statements.get("income_statement") or []
    bs_list = statements.get("balance_sheet") or []
    cf_list = statements.get("cash_flow") or []

    _sheet_from_period_dicts(wb, "Income Statement", is_list)
    _sheet_from_period_dicts(wb, "Balance Sheet", bs_list)
    _sheet_from_period_dicts(wb, "Cash Flow", cf_list)

    if kpis:
        ws = wb.create_sheet(title="KPIs"[:31])
        kpi_keys = [k for k in kpis[0].keys() if k != "period_index"]
        ws.cell(row=1, column=1, value="Period")
        for c, k in enumerate(kpi_keys, start=2):
            ws.cell(row=1, column=c, value=k.replace("_", " ").title())
        for r, row in enumerate(kpis, start=2):
            ws.cell(row=r, column=1, value=row.get("period_index", r - 2))
            for c, k in enumerate(kpi_keys, start=2):
                val = row.get(k)
                if isinstance(val, (int, float)):
                    ws.cell(row=r, column=c, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
