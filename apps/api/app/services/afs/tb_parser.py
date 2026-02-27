"""Parse uploaded Excel/CSV trial balance files into structured account data."""

from __future__ import annotations

import csv
import io
from dataclasses import asdict, dataclass
from typing import Any

import openpyxl


@dataclass
class TBAccount:
    """A single trial balance line item."""
    gl_code: str
    account_name: str
    debit: float
    credit: float
    net: float


@dataclass
class TBParseResult:
    """Result of parsing a trial balance file."""
    accounts: list[TBAccount]
    sheet_name: str | None
    row_count: int
    warnings: list[str]


_GL_HEADERS = {"gl code", "gl_code", "account code", "account_code", "acc code", "code", "gl", "account number", "acc no"}
_NAME_HEADERS = {"account name", "account_name", "description", "name", "account", "account description", "gl description"}
_DEBIT_HEADERS = {"debit", "dr", "debit amount", "debit_amount"}
_CREDIT_HEADERS = {"credit", "cr", "credit amount", "credit_amount"}
_BALANCE_HEADERS = {"balance", "net", "amount", "total", "net amount", "closing balance"}


def _detect_columns(headers: list[str]) -> dict[str, int | None]:
    """Detect column indices from header row using heuristic matching."""
    lower = [h.strip().lower() if h else "" for h in headers]
    result: dict[str, int | None] = {"gl_code": None, "name": None, "debit": None, "credit": None, "balance": None}

    for i, h in enumerate(lower):
        if h in _GL_HEADERS and result["gl_code"] is None:
            result["gl_code"] = i
        elif h in _NAME_HEADERS and result["name"] is None:
            result["name"] = i
        elif h in _DEBIT_HEADERS and result["debit"] is None:
            result["debit"] = i
        elif h in _CREDIT_HEADERS and result["credit"] is None:
            result["credit"] = i
        elif h in _BALANCE_HEADERS and result["balance"] is None:
            result["balance"] = i

    return result


def _safe_float(val: Any) -> float:
    """Convert a cell value to float, defaulting to 0.0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace(",", "").replace(" ", "").strip()
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        return float(cleaned) if cleaned else 0.0
    except (ValueError, TypeError):
        return 0.0


def parse_excel_tb(file_bytes: bytes, filename: str = "tb.xlsx") -> TBParseResult:
    """Parse an Excel trial balance into structured accounts."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        warnings: list[str] = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            for header_idx in range(min(5, len(rows))):
                headers = [str(c) if c is not None else "" for c in rows[header_idx]]
                cols = _detect_columns(headers)

                if cols["name"] is None:
                    continue

                accounts: list[TBAccount] = []
                for row in rows[header_idx + 1:]:
                    cells = list(row)
                    name_val = cells[cols["name"]] if cols["name"] is not None and cols["name"] < len(cells) else None
                    if not name_val or not str(name_val).strip():
                        continue

                    name_str = str(name_val).strip()
                    if name_str.lower().startswith(("total", "sub-total", "subtotal", "grand total")):
                        continue

                    gl = ""
                    if cols["gl_code"] is not None and cols["gl_code"] < len(cells):
                        gl_val = cells[cols["gl_code"]]
                        gl = str(gl_val).strip() if gl_val is not None else ""

                    if cols["debit"] is not None and cols["credit"] is not None:
                        debit = _safe_float(cells[cols["debit"]] if cols["debit"] < len(cells) else None)
                        credit = _safe_float(cells[cols["credit"]] if cols["credit"] < len(cells) else None)
                        net = debit - credit
                    elif cols["balance"] is not None:
                        bal = _safe_float(cells[cols["balance"]] if cols["balance"] < len(cells) else None)
                        debit = bal if bal >= 0 else 0.0
                        credit = abs(bal) if bal < 0 else 0.0
                        net = bal
                    else:
                        warnings.append(f"Row '{name_str}': no numeric columns detected, skipping")
                        continue

                    accounts.append(TBAccount(gl_code=gl, account_name=name_str, debit=debit, credit=credit, net=net))

                if accounts:
                    if cols["gl_code"] is None:
                        warnings.append("GL code column not detected; codes will be empty")
                    return TBParseResult(accounts=accounts, sheet_name=sheet, row_count=len(accounts), warnings=warnings)

        return TBParseResult(accounts=[], sheet_name=None, row_count=0, warnings=["No trial balance data detected in any sheet"])
    finally:
        wb.close()


def parse_csv_tb(file_bytes: bytes) -> TBParseResult:
    """Parse a CSV trial balance into structured accounts."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        return TBParseResult(accounts=[], sheet_name=None, row_count=0, warnings=["CSV file is empty or has only headers"])

    headers = rows[0]
    cols = _detect_columns(headers)
    if cols["name"] is None:
        return TBParseResult(accounts=[], sheet_name=None, row_count=0, warnings=["No account name column detected in CSV"])

    warnings: list[str] = []
    accounts: list[TBAccount] = []
    for row in rows[1:]:
        if cols["name"] >= len(row):
            continue
        name_str = row[cols["name"]].strip()
        if not name_str or name_str.lower().startswith(("total", "sub-total", "subtotal")):
            continue

        gl = row[cols["gl_code"]].strip() if cols["gl_code"] is not None and cols["gl_code"] < len(row) else ""

        if cols["debit"] is not None and cols["credit"] is not None:
            debit = _safe_float(row[cols["debit"]] if cols["debit"] < len(row) else None)
            credit = _safe_float(row[cols["credit"]] if cols["credit"] < len(row) else None)
            net = debit - credit
        elif cols["balance"] is not None:
            bal = _safe_float(row[cols["balance"]] if cols["balance"] < len(row) else None)
            debit = bal if bal >= 0 else 0.0
            credit = abs(bal) if bal < 0 else 0.0
            net = bal
        else:
            continue

        accounts.append(TBAccount(gl_code=gl, account_name=name_str, debit=debit, credit=credit, net=net))

    if cols["gl_code"] is None:
        warnings.append("GL code column not detected; codes will be empty")

    return TBParseResult(accounts=accounts, sheet_name=None, row_count=len(accounts), warnings=warnings)


def tb_accounts_to_json(accounts: list[TBAccount]) -> list[dict[str, Any]]:
    """Convert parsed accounts to JSON-serialisable list."""
    return [asdict(a) for a in accounts]
